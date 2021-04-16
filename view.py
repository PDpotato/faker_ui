# -- coding: utf-8 --

import threading
import time
import asyncio
import requests
import json
from tkinter import *
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import filedialog
from tkinter import ttk
from openpyxl import *
from openpyxl.styles import Font, colors
from generate import Generate
from generate import Type

from mysqlmapper import MysqlMapper as mapper
from history import History as history
from randominfo import RandomInfo as info
from randominfo import Func


class TestView(object):
    def __init__(self, color="#FAFAFA", cursor="hand2", font=("黑体", 11, "bold"), sticky="n" + "s" + "w" + "e",
                 faker=info("zh_CN")):
        self.color = color
        self.view_color = "#666666"
        self.cursor = cursor
        self.font = font
        self.sticky = sticky
        self.faker = faker
        self.window = None
        self.mapper = None
        self.dynamic_generate = Generate()
        self.frame = None
        self.table_frame = None
        self.field_frame = None
        self.previous_btn = None
        self.next_btn = None
        self.save_btn = None
        self.span = None
        self.bar_btn = None
        self.add_btn = None
        self.tip = None
        self.index = 1
        self.is_open_bar = True
        self.default_data = []
        self.content_type = ["None", "multipart/form-data", "application/json"]
        self.methods = {"GET": requests.get, "POST": requests.post, "PUT": requests.put, "DELETE": requests.delete,
                        "OPTIONS": requests.options, "HEAD": requests.head}
        self.history = history()
        self.history_data = {}
        self.auth = None
        self.page = 1
        self.size = 12
        self.host = "localhost"
        self.port = "3306"
        self.username = "root"
        self.password = "123456"
        self.conditions = {Type.ENTITY.value: 1, Type.SERVICE.value: 1, Type.IMPL.value: 1, Type.MAPPER.value: 1,
                           Type.XML.value: 1, Type.CONTROLLER.value: 1}
        self.main_window()

    # 设置打开的窗口居中
    def center_window(self, w, h):
        window = Tk()
        # 窗口居中
        ws = window.winfo_screenwidth() / 2
        hs = window.winfo_screenheight() / 2
        x = ws - w / 2
        y = hs - h / 2
        window.geometry("%dx%d+%d+%d" % (w, h, x, y))
        window.resizable(0, 0)
        window.configure(background=self.color)
        return window

    def main_window(self):
        w = 1400
        h = 800
        self.window = self.center_window(w=w, h=h)
        self.window.title("模拟数据生成工具")
        distance = 50
        operation_width = 150
        top_height = 70
        top_window = PanedWindow(self.window, bg=self.color, height=top_height, width=w, sashwidth=1, orient="vertical")
        top_window.grid(row=0, column=0, columnspan=2)
        font = ("宋体", 15, "bold")
        l1 = Label(top_window, text="", font=font, bg=self.color, width=w)
        self.tip = StringVar()
        l2 = Label(top_window, text="", textvariable=self.tip, font=self.font, bg=self.color, width=w, fg="#99CC33")
        top_window.add(l1, sticky="n")
        top_window.add(l2, sticky="n")
        self.span = ttk.Progressbar(top_window, mode="determinate", orient="horizontal", length=w)
        top_window.add(self.span, sticky="s")
        view_window = PanedWindow(self.window, bg=self.color, height=h - top_height, width=w - operation_width,
                                  sashwidth=1)
        view_window.grid(row=1, column=1)
        operation_window = PanedWindow(self.window, bg=self.color, height=h - top_height, width=operation_width,
                                       sashwidth=1, orient="vertical")
        operation_window.grid(row=1, column=0)
        b1 = Button(operation_window, text="数据库模拟数据", overrelief="ridge", cursor=self.cursor, bg=self.color,
                    relief="groove", font=self.font,
                    command=lambda v=view_window, d=distance: self.sql_frame(v, d, l1, "数据库模拟数据", self.choose_tables))
        b2 = Button(operation_window, text="WEB模拟数据", overrelief="ridge", cursor=self.cursor, bg=self.color,
                    relief="groove", font=self.font,
                    command=lambda v=view_window, d=distance: self.web_frame(v, d, l1, "WEB模拟数据"))
        b3 = Button(operation_window, text="数据库导出EXCEL", overrelief="ridge", cursor=self.cursor, bg=self.color,
                    relief="groove", font=self.font, state="disable",
                    command=lambda v=view_window, d=distance: self.sql_frame(v, d, l1, "数据库导出EXCEL", None))
        b4 = Button(operation_window, text="EXCEL导入数据库", overrelief="ridge", cursor=self.cursor, bg=self.color,
                    relief="groove", font=self.font, state="disable",
                    command=lambda v=view_window, d=distance: self.sql_frame(v, d, l1, "EXCEL导入数据库", None))
        b5 = Button(operation_window, text="生成EXCEL", overrelief="ridge", cursor=self.cursor, bg=self.color,
                    relief="groove", font=self.font,
                    command=lambda v=view_window, d=distance: self.excel_frame(v, d, l1, "生成EXCEL"))
        b6 = Button(operation_window, text="Java代码生成器", overrelief="ridge", cursor=self.cursor, bg=self.color,
                    relief="groove", font=self.font,
                    command=lambda v=view_window, d=distance: self.sql_frame(v, d, l1, "Java代码生成器", self.get_tables))
        l3 = Label(operation_window, text="author:helz", font=self.font, bg=self.color, width=w, fg="#999966")
        operation_window.add(b1, height=60, width=operation_width, sticky="n")
        operation_window.add(b2, height=60, width=operation_width, sticky="n")
        operation_window.add(b3, height=60, width=operation_width, sticky="n")
        operation_window.add(b4, height=60, width=operation_width, sticky="n")
        operation_window.add(b5, height=60, width=operation_width, sticky="n")
        operation_window.add(b6, height=60, width=operation_width, sticky="n")
        operation_window.add(l3, height=60, width=operation_width, sticky="s")
        self.sql_frame(view_window, distance, l1, "数据库模拟数据", self.choose_field)
        self.window.mainloop()

    def excel_frame(self, view_window, distance, label, s):
        label["text"] = s
        if self.frame is not None:
            self.frame.destroy()
        self.frame = Frame(view_window, bg=self.view_color, padx=distance, pady=distance)
        view_window.add(self.frame)
        data = LabelFrame(self.frame, bg=self.view_color, text="设置Excel字段：", font=self.font, pady=10, relief="flat")
        data.grid(row=0, column=0, sticky="nw")
        view_frame = Frame(data, bg=self.view_color)
        view_frame.grid(row=0, column=0, sticky=self.sticky)
        l1 = Label(view_frame, text="表头", bg=self.view_color, font=self.font, relief="groove", height=2, width=20)
        l1.grid(row=0, column=0, sticky=self.sticky)
        l2 = Label(view_frame, text="值类型", bg=self.view_color, font=self.font, relief="groove", height=2, width=20)
        l2.grid(row=0, column=1, sticky=self.sticky)

        operation_frame = Frame(data, bg=self.view_color)
        operation_frame.grid(row=1, column=0, sticky="w")
        self.default_data.clear()
        self.page = 1
        self.size = 12
        self.default_data.clear()
        self.add_btn = Button(operation_frame, text="＋", bg=self.view_color, font=self.font, overrelief="ridge",
                              relief="groove", width=6, cursor=self.cursor, command=lambda: self.add(view_frame))
        self.add_btn.grid(row=0, column=0, sticky="w", pady=2)
        self.previous_btn = Button(operation_frame, text="上页", bg=self.view_color, font=self.font, overrelief="ridge",
                                   relief="groove", width=6, cursor=self.cursor, state="disable",
                                   command=lambda: self.up(view_frame))
        self.previous_btn.grid(row=1, column=0, sticky="w")
        self.next_btn = Button(operation_frame, text="下页", bg=self.view_color, font=self.font, overrelief="ridge",
                               relief="groove", width=6, cursor=self.cursor, state="disable",
                               command=lambda: self.next(view_frame))
        self.next_btn.grid(row=1, column=1, sticky="w")
        self.save_btn = Button(operation_frame, text="导出", bg=self.view_color, font=self.font, overrelief="ridge",
                               relief="groove", width=13, cursor=self.cursor, height=3,
                               command=lambda: self.output())
        self.save_btn.grid(row=2, column=0, sticky="w", columnspan=2, pady=10)

    def output(self):
        # max_num = 1000000
        if len(self.default_data) == 0:
            messagebox.showerror(title="错误", message="请先设置字段值！", parent=self.window)
            return
        column_count = simpledialog.askinteger(parent=self.window, title="输入框", prompt="输入需要导出数据条数：",
                                               initialvalue=1, minvalue=1, maxvalue=1000000)
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", parent=self.window,
                                                filetypes=[("xlsx", ".xlsx")], title="导出Excel")
        wb = Workbook()
        head_font = Font(name="微软雅黑", size=16, bold=True, color=colors.BLACK)
        text_font = Font(name="微软雅黑", size=12, bold=False, color=colors.BLACK)
        start = time.time()
        try:
            # rem = column_count % max_num
            ws = wb.active
            for i in range(len(self.default_data)):
                ws.cell(row=1, column=i + 1, value=self.default_data[i][0].get()).font = head_font
            # _time = column_count // max_num if rem == 0 else column_count // max_num + 1
            # for i in range(_time):
            #     count = rem if i == _time - 1 and rem != 0 else max_num
            self.output_excel(column_count, wb, text_font, filename)
            # wb = load_workbook(filename)
        except Exception as e:
            messagebox.showerror(title="错误", message=e, parent=self.window)
            return
        end = time.time()
        use_time = str(int(round(end * 1000)) - int(round(start * 1000)))
        length = use_time.__len__()
        if length > 3:
            use_time = use_time[0:length - 3] + "," + use_time[length - 3:length]
        messagebox.showinfo(title="提示", message="导出完成！耗时：%sms" % use_time, parent=self.window)

    def output_excel(self, count, wb, text_font, filename):
        end_bar = self.set_span()
        ws = wb.active
        start = ws.max_row
        num = count // 5
        for j in range(count):
            index = 1
            for i in self.default_data:
                if not i[1].get() in self.faker.random_dict:
                    ws.cell(row=start + j + 1, column=index, value=i[1].get()).font = text_font
                    continue
                if self.faker.random_dict.get(i[1].get()) is not None:
                    v = self.faker.random_dict.get(i[1].get())()
                    ws.cell(row=start + j + 1, column=index,
                            value=str(v)).font = text_font
                index += 1
            if j % num == 0 and j > 0:
                end_bar = self.set_span(end_bar, 70 // 5 + end_bar)
        end_bar = self.set_span(end_bar, end_bar + 15)
        wb.save(filename)
        self.set_span(end_bar, 101)
        self.span["value"] = 0
        self.span.update()

    def web_frame(self, view_window, distance, label, s):
        label["text"] = s
        if self.frame is not None:
            self.frame.destroy()
        self.frame = Frame(view_window, bg=self.view_color, padx=distance, pady=distance)
        view_window.add(self.frame)
        method = StringVar()
        method.set(list(self.methods.keys())[0])
        option = OptionMenu(self.frame, method, *list(self.methods.keys()))
        self.set_default(option)
        option["width"] = 10
        option.grid(row=0, column=0)
        url_input = Entry(self.frame, bg=self.view_color, font=self.font, relief="groove", width=65)
        url_input.grid(row=0, column=1, padx=5, sticky=self.sticky)

        f = Frame(self.frame, bg=self.view_color)
        f.grid(row=1, column=0, columnspan=4, sticky=self.sticky, pady=10)

        thread_l = Label(f, text="线程数：", bg=self.view_color, font=self.font, relief="flat", height=2)
        thread_l.grid(row=0, column=0, padx=10, sticky=self.sticky)
        thread_var = IntVar()
        thread_s = Scale(f, bg=self.view_color, font=self.font, length=300, orient="horizontal", to=50, from_=1,
                         highlightbackground=self.view_color, variable=thread_var, cursor=self.cursor)
        thread_s.grid(row=0, column=1, sticky=self.sticky)

        request_l = Label(f, text="请求数：", bg=self.view_color, font=self.font, relief="flat", height=2)
        request_l.grid(row=1, column=0, padx=10, pady=5, sticky=self.sticky)
        request_var = IntVar()
        request_s = Scale(f, bg=self.view_color, font=self.font, length=600, orient="horizontal", to=100,
                          from_=1, highlightbackground=self.view_color, variable=request_var, cursor=self.cursor)
        request_s.grid(row=1, column=1, pady=5, sticky=self.sticky, columnspan=2)
        request_b1 = Button(f, text="×1", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                            width=2, cursor=self.cursor, command=lambda: self.redouble(request_s, 1, request_var))
        request_b1.grid(row=1, column=3, pady=5, sticky="s")
        request_b2 = Button(f, text="×10", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                            width=3, cursor=self.cursor, command=lambda: self.redouble(request_s, 10, request_var))
        request_b2.grid(row=1, column=4, pady=5, sticky="s")
        request_b3 = Button(f, text="×100", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                            width=4, cursor=self.cursor, command=lambda: self.redouble(request_s, 100, request_var))
        request_b3.grid(row=1, column=5, pady=5, sticky="s")
        request_b4 = Button(f, text="×1000", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                            width=5, cursor=self.cursor, command=lambda: self.redouble(request_s, 1000, request_var))
        request_b4.grid(row=1, column=6, pady=5, sticky="s")

        data_f = Frame(self.frame, bg=self.view_color)
        data_f.grid(row=2, column=0, columnspan=4, sticky=self.sticky, pady=10)
        data_l = Label(data_f, text="Content-type：", bg=self.view_color, font=self.font, relief="flat", height=2)
        data_l.grid(row=2, column=0, padx=10, sticky=self.sticky)
        data_var = StringVar()
        data_var.set(self.content_type[0])
        data_option = OptionMenu(data_f, data_var, *self.content_type)
        self.set_default(data_option)
        data_option["width"] = 18
        data_option.grid(row=2, column=1, sticky=self.sticky)
        data_auth = Button(data_f, text="添加Auth", bg=self.view_color, font=self.font, overrelief="ridge",
                           relief="groove",
                           width=10, cursor=self.cursor, command=lambda: self.add_auth())
        data_auth.grid(row=2, column=2, padx=20, sticky=self.sticky)

        data = LabelFrame(self.frame, bg=self.view_color, text="请求参数：", font=self.font, pady=10, relief="flat")
        data.grid(row=3, column=0, columnspan=5, sticky="nw")
        view_frame = Frame(data, bg=self.view_color)
        view_frame.grid(row=1, column=0, columnspan=4, sticky=self.sticky)
        l1 = Label(view_frame, text="key", bg=self.view_color, font=self.font, relief="groove", height=2, width=20)
        l1.grid(row=0, column=0, sticky=self.sticky)
        l2 = Label(view_frame, text="value", bg=self.view_color, font=self.font, relief="groove", height=2, width=20)
        l2.grid(row=0, column=1, sticky=self.sticky)

        operation_frame = Frame(data, bg=self.view_color)
        operation_frame.grid(row=2, column=0, sticky="w")
        self.default_data.clear()
        self.page = 1
        self.size = 12
        self.add_btn = Button(operation_frame, text="＋", bg=self.view_color, font=self.font, overrelief="ridge",
                              relief="groove", width=6, cursor=self.cursor, command=lambda: self.add(view_frame))
        self.add_btn.grid(row=0, column=0, sticky="w", pady=3)
        self.previous_btn = Button(operation_frame, text="上页", bg=self.view_color, font=self.font, overrelief="ridge",
                                   relief="groove", width=6, cursor=self.cursor, state="disable",
                                   command=lambda: self.up(view_frame))
        self.previous_btn.grid(row=1, column=0, sticky="w")
        self.next_btn = Button(operation_frame, text="下页", bg=self.view_color, font=self.font, overrelief="ridge",
                               relief="groove", width=6, cursor=self.cursor, state="disable",
                               command=lambda: self.next(view_frame))
        self.next_btn.grid(row=1, column=1, sticky="w")
        import_btn = Button(operation_frame, text="导入历史配置", bg=self.view_color, font=self.font, overrelief="ridge",
                            relief="groove", width=12, cursor=self.cursor,
                            command=lambda: self.web_import_history(view_frame, method, url_input))
        import_btn.grid(row=1, column=2, sticky="w")
        import_btn.bind("<Enter>", lambda t: self.set_tip(t, "请求成功后会根据请求方式加请求连接地址生成key来存储当前配置，下次即可直接导入"))
        import_btn.bind("<Leave>", lambda t: self.set_tip(t))

        b1 = Button(self.frame, text="请求", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                    width=6, cursor=self.cursor,
                    command=lambda: self.send(method, url_input, thread_var, request_var, data_var))
        b1.grid(row=0, column=2, sticky=self.sticky)
        b2 = Button(self.frame, text="清除", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                    width=6, cursor=self.cursor, command=lambda: self.clear(url_input))
        b2.grid(row=0, column=3, sticky=self.sticky)

    def web_import_history(self, frame, method, url_input):
        url_addr = url_input.get()
        if url_addr == "":
            url_input["highlightbackground"] = "red"
            url_input["highlightcolor"] = "red"
            url_input["highlightthickness"] = 1
            return
        if not url_addr.startswith("http"):
            url_addr = "http://" + url_addr
        key = method.get() + "_" + url_addr
        data = self.history.select_history(key)
        if len(data) != 0:
            d = json.loads(data[0][1])
            for i in range(len(d)):
                if i >= len(self.default_data):
                    key = StringVar()
                    key.set(d[i][0])
                    value = StringVar()
                    value.set(d[i][1])
                    self.default_data.append([key, value])
                else:
                    self.default_data[i][0].set(d[i][0])
                    self.default_data[i][1].set(d[i][1])
            self.change_web_data(frame)
        else:
            messagebox.showwarning(parent=self.window, title="警告", message="该请求地址无历史记录")

    def add_auth(self):
        auth_window = self.center_window(300, 280)
        auth_window.title("添加Auth")
        auth_window.configure(bg=self.view_color)
        text = Text(auth_window, bg=self.color, width=33, height=15, font=self.font)
        text.grid(row=0, column=0, sticky=self.sticky)
        btn = Button(auth_window, text="添加", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                     width=6, cursor=self.cursor, command=lambda v=text.get(1.0, END): self.set_auth(v, auth_window))
        btn.grid(row=1, column=0, pady=5)

    def set_auth(self, value, window):
        self.auth = value
        window.destroy()

    def up(self, frame):
        self.page -= 1
        self.change_web_data(frame)

    def next(self, frame):
        self.page += 1
        self.change_web_data(frame)

    def change_web_data(self, frame):
        self.index = 1
        begin = (self.page - 1) * self.size
        end = self.page * self.size if self.page * self.size < len(self.default_data) else len(self.default_data)
        values = self.default_data[begin: end]
        modules = list(frame.children.values())
        for i in range(modules.__len__()):
            if i >= 2:
                modules[i].destroy()
        for j in values:
            e = Entry(frame, bg=self.view_color, font=self.font, relief="groove", textvariable=j[0])
            e.grid(row=self.index, column=0, sticky=self.sticky)
            option = OptionMenu(frame, j[1], *list(self.faker.random_dict.keys()))
            self.set_default(option)
            option.grid(row=self.index, column=1, sticky=self.sticky)
            data_b1 = Button(frame, text="自定义", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                             width=6, cursor=self.cursor, command=lambda choose=j[1]: self.child_add(choose))
            data_b1.grid(row=self.index, column=2, sticky=self.sticky)
            data_b2 = Button(frame, text="×", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                             width=2, cursor=self.cursor, fg="#CC3333")
            data_b2.grid(row=self.index, column=3, sticky=self.sticky)
            data_b2.bind("<Button-1>", lambda f=None, v=j: self.child_delete(f, frame, v))
            self.index += 1
        max_page = self.max_page(self.default_data)
        # 此处为了容错
        self.save_btn = None
        self.change_turning(max_page)
        if self.page < max_page:
            self.add_btn.destroy()
        else:
            if self.add_btn is not None:
                self.add_btn.destroy()
            self.add_btn = Button(self.previous_btn.master, text="＋", bg=self.view_color, font=self.font,
                                  overrelief="ridge", relief="groove", width=6, cursor=self.cursor,
                                  command=lambda: self.add(frame))
            self.add_btn.grid(row=0, column=0, sticky="w")

    def add(self, frame):
        if self.default_data.__len__() > 0 and self.default_data.__len__() % self.size == 0:
            modules = list(frame.children.values())
            for i in range(modules.__len__()):
                if i >= 2:
                    modules[i].destroy()
            self.index = 1
            self.page += 1
            # 此处为了容错
            self.save_btn = None
            self.change_turning(self.max_page(self.default_data))
        inner = []
        input_str = StringVar()
        e = Entry(frame, bg=self.view_color, font=self.font, relief="groove", textvariable=input_str)
        e.grid(row=self.index, column=0, sticky=self.sticky)
        inner.insert(0, input_str)
        choose = StringVar()
        inner.insert(1, choose)
        choose.set(Func.WORD.value)
        self.default_data.append(inner)
        option = OptionMenu(frame, choose, *list(self.faker.random_dict.keys()))
        self.set_default(option)
        option.grid(row=self.index, column=1, sticky=self.sticky)
        data_b1 = Button(frame, text="自定义", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                         width=6, cursor=self.cursor, command=lambda: self.child_add(choose))
        data_b1.grid(row=self.index, column=2, sticky=self.sticky)
        data_b2 = Button(frame, text="×", bg=self.view_color, font=self.font, overrelief="ridge", relief="groove",
                         width=2, cursor=self.cursor, fg="#CC3333")
        data_b2.grid(row=self.index, column=3, sticky=self.sticky)
        data_b2.bind("<Button-1>", lambda f=None, v=inner: self.child_delete(f, frame, v))
        self.index += 1

    def child_delete(self, f, frame, value):
        self.default_data.remove(value)
        widgets = list(frame.children.values())
        index = f.widget.grid_info().get("row")
        length = len(self.default_data)
        for i in widgets:
            if i.grid_info().get("row") == index:
                i.destroy()
        if length > 0 and length % self.size == 0 and index + (self.page - 1) * self.size == (length + 1):
            self.up(frame)

    def child_add(self, choose):
        a = simpledialog.askstring(title="输入", prompt="请输入默认值")
        if a is not None:
            choose.set(str(a))

    def redouble(self, s, num, var):
        n = var.get() / s["to"]
        s["to"] = 100 * num
        var.set(int(n * s["to"]))

    def send(self, method, url, thread_var, request_var, data_var):
        url_adr = url.get()
        if url_adr == "":
            url["highlightbackground"] = "red"
            url["highlightcolor"] = "red"
            url["highlightthickness"] = 1
            return
        data_var = data_var.get().strip()
        func = self.methods.get(method.get())
        if not url_adr.startswith("http"):
            url_adr = "http://" + url.get()
        start = time.time()
        count = request_var.get() // thread_var.get()
        header = {"Authorization": self.auth}
        if data_var != self.content_type[0]:
            header["Content-Type"] = data_var
        try:
            for i in range(count):
                self.refresh(func, url_adr, header, data_var)
        except Exception as e:
            messagebox.showerror(title="警告", message=e, parent=self.window)
            return
        for j in range(thread_var.get() - 1):
            if j == thread_var.get() - 1:
                count = count + request_var.get() % thread_var.get()
            t = threading.Thread(
                target=lambda: [self.refresh(func, url.get(), header, data_var) for k in range(count)])
            t.setDaemon(True)
            t.start()
        end = time.time()
        use_time = str(int(round(end * 1000)) - int(round(start * 1000)))
        message = "任务开启成功！执行线程数%s，执行次数%s，耗时:%sms" % (thread_var.get(), request_var.get(), use_time)
        messagebox.showinfo(title="提示", message=message, parent=self.window)
        key = method.get() + "_" + url_adr
        data = []
        for i in self.default_data:
            d = []
            for j in i:
                d.append(j.get())
            data.append(d)
        self.history.update_history(key, json.dumps(data, ensure_ascii=False))

    def refresh(self, func, url, header, data):
        body = {}
        for i in self.default_data:
            if not i[1].get() in self.faker.random_dict:
                body[i[0].get()] = i[1].get()
                continue
            if self.faker.random_dict.get(i[1].get()) is not None:
                body[i[0].get()] = self.faker.random_dict.get(i[1].get())()
        if data == self.content_type[0]:
            resp = func(url, params=body, headers=header)
        elif data == self.content_type[1]:
            resp = func(url, data=body, headers=header)
        elif data == self.content_type[2]:
            resp = func(url, data=json.dumps(body), headers=header)
        if resp.status_code != 200:
            raise Exception(resp.content)

    def clear(self, url_input):
        url_input.delete(0, END)

    def sql_frame(self, view_window, distance, label, s, func):
        label["text"] = s
        if self.frame is not None:
            self.frame.destroy()
        self.frame = Frame(view_window, bg=self.view_color, padx=distance, pady=distance)
        view_window.add(self.frame)
        x = 15
        y = 5
        host = StringVar()
        host.set(self.host)
        port = StringVar()
        port.set(self.port)
        user = StringVar()
        user.set(self.username)
        password = StringVar()
        password.set(self.password)
        lf = LabelFrame(self.frame, text="请输入MYSQL连接信息", bg=self.view_color, font=self.font, pady=5, relief="sunken")
        lf.grid(row=0, column=0)
        Label(lf, text="主机:", bg=self.view_color, font=self.font, anchor="e", width=8).grid(row=0, column=0)
        Entry(lf, bg=self.view_color, textvariable=host, font=self.font).grid(row=0, column=1, padx=x, pady=y)
        Label(lf, text="端口:", bg=self.view_color, font=self.font, anchor="e", width=8).grid(row=1, column=0)
        Entry(lf, bg=self.view_color, textvariable=port, font=self.font).grid(row=1, column=1, padx=x, pady=y)
        Label(lf, text="用户名:", bg=self.view_color, font=self.font, anchor="e", width=8).grid(row=2, column=0)
        Entry(lf, bg=self.view_color, textvariable=user, font=self.font).grid(row=2, column=1, padx=x, pady=y)
        Label(lf, text="密码:", bg=self.view_color, font=self.font, anchor="e", width=8).grid(row=3, column=0)
        Entry(lf, bg=self.view_color, show="*", textvariable=password, font=self.font).grid(row=3, column=1, padx=x,
                                                                                            pady=y)
        Button(lf, text="连接", bg=self.view_color, font=self.font, overrelief="ridge", cursor=self.cursor, width=10,
               relief="groove", command=lambda: self.connect(host, port, user, password, lf, func)). \
            grid(row=4, column=1)

    def connect(self, host, port, user, password, frame, func):
        if host.get() == "":
            messagebox.showwarning(parent=self.window, title="警告", message="主机不能为空")
            return
        if port.get() == "":
            messagebox.showwarning(parent=self.window, title="警告", message="端口不能为空")
            return
        if user.get() == "":
            messagebox.showwarning(parent=self.window, title="警告", message="用户名不能为空")
            return
        if password.get() == "":
            messagebox.showwarning(parent=self.window, title="警告", message="密码不能为空")
            return
        try:
            self.mapper = mapper(host.get(), port.get(), user.get(), password.get())
        except Exception as e:
            messagebox.showerror(parent=self.window, title="错误", message=str(e))
            return
        databases = []
        messagebox.showinfo(parent=self.window, title="正确", message="连接成功")
        self.host = host.get()
        self.port = port.get()
        self.username = user.get()
        self.password = password.get()
        for i in self.mapper.show_databases():
            databases.append(i[0])
        frame.destroy()
        lf = LabelFrame(self.frame, text="请选择数据库", bg=self.view_color, font=self.font, pady=5, relief="flat")
        lf.grid(row=0, column=0, sticky="n")
        database = StringVar()
        option = OptionMenu(lf, database, *databases,
                            command=lambda data=database.get(): func(data))
        option.grid(row=0, column=0)
        self.set_default(option)

    def choose_tables(self, database):
        if self.table_frame is not None:
            self.table_frame.destroy()
        if self.field_frame is not None:
            self.field_frame.destroy()
        tables = []
        for i in self.mapper.show_tables(database):
            tables.append(i[0])
        self.table_frame = LabelFrame(self.frame, text="请选择表", bg=self.view_color, font=self.font, pady=5,
                                      relief="flat")
        self.table_frame.grid(row=0, column=1, sticky="n")
        table = StringVar()
        option = OptionMenu(self.table_frame, table, *tables,
                            command=lambda t=table.get(): self.choose_field(t, database))
        option.grid(row=0, column=0)
        self.set_default(option)

    def get_tables(self, database):
        if self.table_frame is not None:
            self.table_frame.destroy()
            self.table_frame = None
        if self.field_frame is not None:
            self.field_frame.destroy()
            self.field_frame = None
        self.table_frame = LabelFrame(self.frame, text="请选择表", bg=self.view_color, font=self.font, pady=5,
                                      relief="groove")
        self.table_frame.grid(row=0, column=1, sticky="n")
        st = Scrollbar(self.table_frame, bg=self.view_color, relief="groove", cursor=self.cursor, orient="horizontal")
        sl = Scrollbar(self.table_frame, bg=self.view_color, relief="groove", cursor=self.cursor)
        st.pack(side="bottom", fill="x")
        sl.pack(side="right", fill="y")
        lb = Listbox(self.table_frame, bg=self.view_color, font=self.font, selectmode="extended", height=20,
                     cursor=self.cursor, width=30, relief="groove", yscrollcommand=sl.set, xscrollcommand=st.set)
        lb.pack(side="left", fill="both")
        sl.config(command=lb.yview)
        st.config(command=lb.xview)
        for i in self.mapper.show_tables(database):
            lb.insert("end", i[0])
        lb.bind("<Button-1>", lambda f=None: self.operation(f))

    def operation(self, f):
        lb = f.widget
        if self.field_frame is None:
            self.field_frame = LabelFrame(self.frame, text="操作", bg=self.view_color, font=self.font, pady=5, padx=5,
                                          relief="groove")
            self.field_frame.grid(row=0, column=2, sticky="n")
            src = StringVar()
            file_src = Entry(self.field_frame, bg=self.view_color, font=self.font, relief="groove", width=65,
                             textvariable=src)
            file_src.grid(row=0, column=0, sticky="nsew")
            file_btn = Button(self.field_frame, text="选择地址", bg=self.view_color, font=self.font, overrelief="ridge",
                              cursor=self.cursor, width=10, relief="groove",
                              command=lambda: src.set(filedialog.askdirectory()))
            file_btn.grid(row=0, column=1, sticky="nsew")
            package = LabelFrame(self.field_frame, text="包路径", bg=self.view_color, font=self.font, pady=5, padx=5,
                                 relief="groove")
            package.grid(row=1, column=0, sticky="nsew", columnspan=2, pady=10)
            p_src = StringVar()
            p_src.set("com.hlz.demo")
            package_src = Entry(package, bg=self.view_color, font=self.font, relief="groove", width=65,
                                textvariable=p_src)
            package_src.grid(row=0, column=0, sticky="nsew")
            user_info = LabelFrame(self.field_frame, text="author", bg=self.view_color, font=self.font, pady=5, padx=5,
                                   relief="groove")
            user_info.grid(row=2, column=0, sticky="nsew", columnspan=2, pady=10)
            user_src = StringVar()
            user_src.set("hlz")
            package_src = Entry(user_info, bg=self.view_color, font=self.font, relief="groove", width=65,
                                textvariable=user_src)
            package_src.grid(row=0, column=0, sticky="nsew")
            choose = LabelFrame(self.field_frame, text="生成选项", bg=self.view_color, font=self.font, pady=5, padx=5,
                                relief="groove")
            choose.grid(row=3, column=0, sticky="nsew", columnspan=2, pady=10)

            for i in range(len(self.conditions)):
                v1 = IntVar()
                v1.set(1)
                key = list(self.conditions.keys())[i]
                c1 = Checkbutton(choose, text=key, variable=v1, bg=self.view_color, anchor="w",
                                 font=self.font, selectcolor=self.view_color, activebackground=self.view_color,
                                 width=15, command=lambda v=v1, k=key: self.update_condition(k, v.get()))
                current = i // 4
                size = i % 4
                c1.grid(row=current, column=size, pady=10, sticky="w")
            btn = Button(self.field_frame, text="生成", bg=self.view_color, font=self.font, overrelief="ridge",
                         cursor=self.cursor, width=10, relief="groove", height=3,
                         command=lambda: self.generate(lb, src, file_src, p_src, package_src, user_src))
            btn.grid(row=4, column=0, sticky="w", pady=5)

    def update_condition(self, key, value):
        self.conditions[key] = value

    def generate(self, lb, src, file_src, p_src, package_src, user_src):
        if src.get() == "":
            file_src["highlightbackground"] = "red"
            file_src["highlightcolor"] = "red"
            file_src["highlightthickness"] = 1
            return
        else:
            file_src["highlightthickness"] = 0
        if p_src.get() == "":
            package_src["highlightbackground"] = "red"
            package_src["highlightcolor"] = "red"
            package_src["highlightthickness"] = 1
            return
        else:
            package_src["highlightthickness"] = 0
        package = p_src.get()
        for i in lb.curselection():
            table = lb.get(i)
            fields = self.mapper.show_field(table)
            table_status = self.mapper.show_table_status(table)
            user = user_src.get()
            if self.conditions.get(Type.ENTITY.value) == 1:
                self.dynamic_generate.generate_entity(package, src.get(), table, fields, table_status[0][17], user)
            if self.conditions.get(Type.MAPPER.value) == 1:
                self.dynamic_generate.generate_mapper(package, src.get(), table, fields, table_status[0][17], user)
            if self.conditions.get(Type.SERVICE.value) == 1:
                self.dynamic_generate.generate_service(package, src.get(), table, fields, table_status[0][17], user)
            if self.conditions.get(Type.IMPL.value) == 1:
                self.dynamic_generate.generate_impl(package, src.get(), table, fields, table_status[0][17], user)
            if self.conditions.get(Type.CONTROLLER.value) == 1:
                self.dynamic_generate.generate_controller(package, src.get(), table, fields, table_status[0][17], user)
            if self.conditions.get(Type.XML.value) == 1:
                self.dynamic_generate.generate_xml(package, src.get(), table, fields, table_status[0][17], user)
        messagebox.showinfo(parent=self.window, title="正确", message="生成完成")

    def choose_field(self, table, database):
        self.page = 1
        self.size = 15
        if self.field_frame is not None:
            self.field_frame.destroy()
        fields = self.mapper.show_field(table)
        self.default_data.clear()
        for i in range(fields.__len__()):
            self.default_data.append([None, None, None])
        self.field_frame = LabelFrame(self.frame, text="请选择数据", bg=self.view_color, font=self.font, pady=5, padx=5)
        self.field_frame.grid(row=0, column=2, sticky="n")
        title_l = Label(self.field_frame, text="字段名", bg=self.view_color, font=self.font, width=12, height=2)
        title_l.grid(row=0, column=0)
        title_l = Label(self.field_frame, text="字段类型", bg=self.view_color, font=self.font, width=12, height=2)
        title_l.grid(row=0, column=1)
        title_l = Label(self.field_frame, text="描述", bg=self.view_color, font=self.font, width=20, height=2)
        title_l.grid(row=0, column=2)
        title_l = Label(self.field_frame, text="能否为null", bg=self.view_color, font=self.font, width=10, height=2)
        title_l.grid(row=0, column=3)
        title_l = Label(self.field_frame, text="数据类型", bg=self.view_color, font=self.font, width=12, height=2)
        title_l.grid(row=0, column=4)
        title_l = Label(self.field_frame, text="设置", bg=self.view_color, font=self.font, width=12, height=2)
        title_l.grid(row=0, column=5)
        module = []
        for i in range(self.size):
            module.append([None, None, None, None, None, None])
        self.change_data(fields, module)
        operation_frame = Frame(self.field_frame, bg=self.view_color)
        operation_frame.grid(row=self.size + 1, column=1, columnspan=3, sticky="w")
        self.previous_btn = Button(operation_frame, text="上页", bg=self.view_color, font=self.font, overrelief="ridge",
                                   state="disabled", cursor=self.cursor, width=5, relief="groove",
                                   command=lambda: self.previous_page(fields, module))
        self.previous_btn.grid(row=0, column=0, sticky=self.sticky)
        self.next_btn = Button(operation_frame, text="下页", bg=self.view_color, font=self.font, overrelief="ridge",
                               state="disabled", cursor=self.cursor, width=5, relief="groove",
                               command=lambda: self.next_page(fields, module))
        self.next_btn.grid(row=0, column=1, sticky=self.sticky)
        self.save_btn = Button(operation_frame, text="插入数据", bg=self.view_color, font=self.font, overrelief="ridge",
                               state="disabled", cursor=self.cursor, width=10, relief="groove",
                               command=lambda: self.int_dialog(table, database))
        self.save_btn.grid(row=0, column=2, sticky=self.sticky)
        self.save_btn.bind("<Enter>", lambda t: self.set_tip(t, "导入历史记录或浏览完全部字段才允许插入数据"))
        self.save_btn.bind("<Leave>", lambda t: self.set_tip(t))
        self.bar_btn = Button(operation_frame, text="导入历史记录", bg=self.view_color, font=self.font, overrelief="ridge",
                              cursor=self.cursor, width=12, relief="groove",
                              command=lambda: self.import_history(database + "_" + table, fields, module))
        self.bar_btn.grid(row=0, column=3, sticky=self.sticky)
        self.bar_btn.bind("<Enter>", lambda t: self.set_tip(t, "导入该表已成功插入的最新记录"))
        self.bar_btn.bind("<Leave>", lambda t: self.set_tip(t))
        if fields.__len__() > self.size:
            self.next_btn["state"] = "normal"
        if self.page >= self.max_page(fields):
            self.save_btn["state"] = "normal"

    def set_tip(self, event, value=""):
        self.tip.set(value)

    def import_history(self, key, fields, module):
        data = self.history.select_history(key)
        if len(data) != 0:
            data = json.loads(data[0][1])
            if len(data) != len(fields):
                messagebox.showerror(parent=self.window, title="错误", message="该表结构已发生改变，无法导入")
                return
            for i in range(len(data)):
                if self.default_data[i][1] is not None:
                    self.default_data[i][1].set(data[i])
                else:
                    field = StringVar()
                    field.set(data[i])
                    self.default_data[i][0] = fields[i][0]
                    self.default_data[i][1] = field
                    s = fields[i][1]
                    if s.find("(") != -1:
                        s = s[0:s.find("(")]
                    self.default_data[i][2] = s
            self.change_data(fields, module)
            messagebox.showinfo(parent=self.window, title="提示", message="导入成功")
            self.save_btn["state"] = "normal"
        else:
            messagebox.showwarning(parent=self.window, title="警告", message="该表无历史记录")

    def next_page(self, fields, module):
        self.page += 1
        self.change_data(fields, module)
        self.change_turning(self.max_page(fields))

    def previous_page(self, fields, module):
        self.page -= 1
        self.change_data(fields, module)
        self.change_turning(self.max_page(fields))

    def change_turning(self, max_page):
        self.previous_btn["state"] = "normal"
        self.next_btn["state"] = "normal"
        if self.page <= 1:
            self.previous_btn["state"] = "disabled"
        if self.page >= max_page:
            self.next_btn["state"] = "disabled"
        if self.page >= max_page and self.save_btn is not None:
            self.save_btn["state"] = "normal"

    def max_page(self, fields):
        total = fields.__len__()
        return total // self.size if total % self.size == 0 else total // self.size + 1

    def change_data(self, fields, module):
        label_style = "solid"
        font_length = 11
        count = (self.page - 1) * self.size
        for i in range(self.size):
            index = count + i
            inner_module = [None, None, None, None, None, None]
            if index < fields.__len__():
                data = []
                s = fields[index][0]
                data.insert(0, s)
                if s.__len__() > font_length:
                    s = s[0:font_length] + "..."
                l1 = module[i][0]
                if l1 is None:
                    l1 = Label(self.field_frame, text=s, bg=self.view_color, font=self.font, width=12, anchor="w",
                               height=2, relief=label_style)
                    l1.grid(row=i + 1, column=0, sticky=self.sticky, pady=1, padx=1)
                else:
                    l1["text"] = s
                    l1["relief"] = label_style
                l1.bind("<Enter>", lambda t, value=fields[index][0]: self.set_tip(t, value))
                l1.bind("<Leave>", lambda t: self.set_tip(t))
                inner_module[0] = l1
                s1 = fields[index][1]
                if s1.find("(") != -1:
                    s1 = s1[0:s1.find("(")]
                l2 = module[i][1]
                if l2 is None:
                    l2 = Label(self.field_frame, text=s1, bg=self.view_color, font=self.font, width=12, anchor="w",
                               relief=label_style)
                    l2.grid(row=i + 1, column=1, sticky=self.sticky, pady=1, padx=1)
                else:
                    l2["text"] = s1
                    l2["relief"] = label_style
                l2.bind("<Enter>", lambda t, value=fields[index][1]: self.set_tip(t, value))
                l2.bind("<Leave>", lambda t: self.set_tip(t))
                inner_module[1] = l2
                s2 = fields[index][8]
                if s2.__len__() > font_length:
                    s2 = s2[0:font_length] + "..."
                l3 = module[i][4]
                if l3 is None:
                    l3 = Label(self.field_frame, text=s2, bg=self.view_color, font=self.font, width=20, anchor="w",
                               height=2, relief=label_style)
                    l3.grid(row=i + 1, column=2, sticky=self.sticky, pady=1, padx=1)
                else:
                    l3["text"] = s2
                    l3["relief"] = label_style
                l3.bind("<Enter>", lambda t, value=fields[index][8]: self.set_tip(t, value))
                l3.bind("<Leave>", lambda t: self.set_tip(t))
                inner_module[4] = l3
                s3 = fields[index][3]
                l4 = module[i][5]
                if l4 is None:
                    l4 = Label(self.field_frame, text=s3, bg=self.view_color, font=self.font, width=12, anchor="w",
                               height=2, relief=label_style)
                    l4.grid(row=i + 1, column=3, sticky=self.sticky, pady=1, padx=1)
                else:
                    l4["text"] = s3
                    l4["relief"] = label_style
                inner_module[5] = l4
                field = StringVar()
                data.insert(1, field)
                data.insert(2, s1)
                b = module[i][3]
                if b is None:
                    b = Button(self.field_frame, text="自定义", bg=self.view_color, font=self.font, overrelief="ridge",
                               cursor=self.cursor, width=10, relief="groove",
                               command=lambda v=field: self.open_dialog(v))
                    b.grid(row=i + 1, column=5, sticky=self.sticky)
                else:
                    b["command"] = lambda v=field: self.open_dialog(v)
                b.bind("<Enter>", lambda t, value="设置自定义后，将固定该值，不会生成随机数据替换": self.set_tip(t, value))
                b.bind("<Leave>", lambda t: self.set_tip(t))
                inner_module[3] = b
                if self.default_data[index][1] is not None:
                    field.set(self.default_data[index][1].get())
                else:
                    if s1 == "boolean" or s1 == "tinyint":
                        field.set(Func.BOOLEAN.value)
                    elif fields[index][0].lower().find("id") != -1 and s1 != "int":
                        field.set(Func.SNOW_ID.value)
                    elif fields[index][0].lower().find("id") != -1:
                        field.set(Func.NONE.value)
                    elif s1 == "int":
                        field.set(Func.RANDOM_NUMBER.value)
                    elif s1 == "datetime":
                        field.set(Func.PAST_DATETIME.value)
                    elif s1 == "varchar":
                        field.set(Func.WORD.value)
                    elif s1 == "boolean":
                        field.set(Func.BOOLEAN.value)
                    elif s1 == "json":
                        field.set(Func.SIMPLE_PROFILE.value)
                    elif s1 == "":
                        field.set(Func.NONE.value)
                        b["state"] = "disabled"
                    else:
                        field.set(Func.WORD.value)
                option = module[i][2]
                if option is None:
                    option = OptionMenu(self.field_frame, field, *self.faker.random_dict.keys())
                    option.grid(row=i + 1, column=4, sticky="w")
                    self.set_default(option)
                    option["width"] = 10
                else:
                    option.destroy()
                    option = OptionMenu(self.field_frame, field, *self.faker.random_dict.keys())
                    option.grid(row=i + 1, column=4, sticky="w")
                    self.set_default(option)
                    option["width"] = 10
                inner_module[2] = option
                self.default_data[index] = data
            else:
                empty = ""
                l1 = module[i][0]
                if l1 is None:
                    l1 = Label(self.field_frame, text=empty, bg=self.view_color, font=self.font, width=12, height=2)
                    l1.grid(row=i + 1, column=0, sticky=self.sticky)
                else:
                    l1["text"] = empty
                    l1["relief"] = "flat"
                inner_module.insert(0, l1)
                l1.bind("<Enter>", lambda t, value=empty: self.set_tip(t, value))
                l1.bind("<Leave>", lambda t: self.set_tip(t))
                l2 = module[i][1]
                if l2 is None:
                    l2 = Label(self.field_frame, text=empty, bg=self.view_color, font=self.font, width=12)
                    l2.grid(row=i + 1, column=1, sticky=self.sticky)
                else:
                    l2["text"] = empty
                    l2["relief"] = "flat"
                inner_module.insert(1, l2)
                l2.bind("<Enter>", lambda t, value=empty: self.set_tip(t, value))
                l2.bind("<Leave>", lambda t: self.set_tip(t))
                l3 = module[i][4]
                if l3 is None:
                    l3 = Label(self.field_frame, text=empty, bg=self.view_color, font=self.font, width=12)
                    l3.grid(row=i + 1, column=2, sticky=self.sticky)
                else:
                    l3["text"] = empty
                    l3["relief"] = "flat"
                inner_module.insert(4, l3)
                l3.bind("<Enter>", lambda t, value=empty: self.set_tip(t, value))
                l3.bind("<Leave>", lambda t: self.set_tip(t))
                l4 = module[i][5]
                if l4 is None:
                    l4 = Label(self.field_frame, text=empty, bg=self.view_color, font=self.font, width=12)
                    l4.grid(row=i + 1, column=3, sticky=self.sticky)
                else:
                    l4["text"] = empty
                    l4["relief"] = "flat"
                inner_module.insert(5, l4)
                l4.bind("<Enter>", lambda t, value=empty: self.set_tip(t, value))
                l4.bind("<Leave>", lambda t: self.set_tip(t))
                inner_module.insert(2, None)
                inner_module.insert(3, None)
                if module[i][2] is not None:
                    module[i][2].destroy()
                if module[i][3] is not None:
                    module[i][3].destroy()
            module[i] = inner_module

    def open_dialog(self, field):
        a = simpledialog.askstring(parent=self.window, title="输入框", prompt="输入默认值", initialvalue=field.get())
        if a is not None:
            field.set(a)

    def set_span(self, start=0, end=10):
        if self.is_open_bar:
            for i in range(start, end):
                self.span["value"] = i
                self.span.update()
                time.sleep(0.005)
        return end

    def int_dialog(self, table, database):
        a = simpledialog.askinteger(parent=self.window, title="输入框", prompt="输入需要插入数据条数", initialvalue=1, minvalue=1,
                                    maxvalue=1000000)
        data = []
        data_history = []
        if a is not None:
            for i in self.default_data:
                inner = (i[0], i[1].get(), i[2])
                data_history.append(i[1].get())
                data.append(inner)
            try:
                # t = threading.Thread(target=self.set_span, args=(a,), name="span_thread")
                # t.start()
                end_bar = self.set_span()
                start = time.time()
                count = 2000
                num = a // count
                for i in range(num):
                    self.faker.insert_data(data, count, table, self.mapper)
                    end_bar = self.set_span(end_bar, 75 // (num - i) + 10)
                count = a % count
                end_bar = self.set_span(end_bar, 95)
                if count > 0:
                    self.faker.insert_data(data, count, table, self.mapper)
                self.mapper.conn.commit()
                end = time.time()
                self.set_span(end_bar, 101)
                use_time = str(int(round(end * 1000)) - int(round(start * 1000)))
                length = use_time.__len__()
                if length > 3:
                    use_time = use_time[0:length - 3] + "," + use_time[length - 3:length]
                messagebox.showinfo(parent=self.window, title="正确", message="数据添加成功(耗时：" + use_time + "毫秒)")
                self.history.update_history(database + "_" + table, json.dumps(data_history, ensure_ascii=False))
            except Exception as e:
                messagebox.showerror(parent=self.window, title="错误", message=str(e))
            self.span["value"] = 0
            self.span.update()

    def set_default(self, option):
        option["bg"] = self.view_color
        option["width"] = 15
        option["cursor"] = self.cursor
        option["font"] = self.font
        option["anchor"] = "w"
        option["relief"] = "flat"
        option["highlightbackground"] = self.color
        option["highlightthickness"] = 1


if __name__ == '__main__':
    view = TestView()
